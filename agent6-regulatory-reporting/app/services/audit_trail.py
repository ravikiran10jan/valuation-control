"""Audit Trail service for SOX compliance.

Provides immutable logging of all valuation-related events.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Optional
import uuid

from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
import structlog

from app.core.config import settings
from app.models.postgres import AuditEvent
from app.models.schemas import (
    AuditEventCreate,
    AuditEventOut,
    AuditEventType,
    AuditReportOut,
    AuditTrailQuery,
)

log = structlog.get_logger()


class AuditTrail:
    """Manage audit trail for regulatory compliance.

    SOX compliance requires:
    - Who did what, when
    - Complete history
    - Tamper-proof logging
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    async def log_audit_event(
        self,
        event_type: AuditEventType,
        user: str,
        details: dict,
        ip_address: Optional[str] = None,
    ) -> AuditEventOut:
        """Log a valuation-related event for audit.

        Args:
            event_type: Type of event (VALUATION_RUN, MARK_ADJUSTMENT, etc.).
            user: User who performed the action.
            details: Event details dictionary.
            ip_address: Optional client IP address.

        Returns:
            Created AuditEventOut.
        """
        audit_event = AuditEvent(
            event_id=uuid.uuid4(),
            event_type=event_type.value,
            user=user,
            timestamp=datetime.utcnow(),
            details=details,
            ip_address=ip_address,
        )

        self.db.add(audit_event)
        await self.db.commit()
        await self.db.refresh(audit_event)

        log.info(
            "audit_event_logged",
            event_id=str(audit_event.event_id),
            event_type=event_type.value,
            user=user,
        )

        return AuditEventOut(
            event_id=str(audit_event.event_id),
            event_type=AuditEventType(audit_event.event_type),
            user=audit_event.user,
            timestamp=audit_event.timestamp,
            details=audit_event.details,
            ip_address=audit_event.ip_address,
        )

    async def get_audit_event(self, event_id: str) -> Optional[AuditEventOut]:
        """Get a single audit event by ID.

        Args:
            event_id: The event UUID.

        Returns:
            AuditEventOut or None if not found.
        """
        try:
            event_uuid = uuid.UUID(event_id)
        except ValueError:
            return None

        stmt = select(AuditEvent).where(AuditEvent.event_id == event_uuid)
        result = await self.db.execute(stmt)
        event = result.scalar_one_or_none()

        if not event:
            return None

        return AuditEventOut(
            event_id=str(event.event_id),
            event_type=AuditEventType(event.event_type),
            user=event.user,
            timestamp=event.timestamp,
            details=event.details,
            ip_address=event.ip_address,
        )

    async def query_audit_events(
        self, query: AuditTrailQuery
    ) -> list[AuditEventOut]:
        """Query audit events with filters.

        Args:
            query: Query parameters.

        Returns:
            List of matching AuditEventOut.
        """
        stmt = select(AuditEvent)

        filters = []

        # Date range filter
        start_datetime = datetime.combine(query.start_date, datetime.min.time())
        end_datetime = datetime.combine(query.end_date, datetime.max.time())
        filters.append(AuditEvent.timestamp.between(start_datetime, end_datetime))

        # Event type filter
        if query.event_type:
            filters.append(AuditEvent.event_type == query.event_type.value)

        # User filter
        if query.user:
            filters.append(AuditEvent.user == query.user)

        if filters:
            stmt = stmt.where(and_(*filters))

        stmt = (
            stmt.order_by(AuditEvent.timestamp.desc())
            .limit(query.limit)
            .offset(query.offset)
        )

        result = await self.db.execute(stmt)
        events = result.scalars().all()

        return [
            AuditEventOut(
                event_id=str(e.event_id),
                event_type=AuditEventType(e.event_type),
                user=e.user,
                timestamp=e.timestamp,
                details=e.details,
                ip_address=e.ip_address,
            )
            for e in events
        ]

    async def generate_audit_report(
        self, start_date: date, end_date: date
    ) -> AuditReportOut:
        """Generate audit report for external auditors.

        Shows all valuation activities during the period.

        Args:
            start_date: Report period start.
            end_date: Report period end.

        Returns:
            AuditReportOut with event summary and details.
        """
        log.info(
            "generating_audit_report",
            start_date=str(start_date),
            end_date=str(end_date),
        )

        # Query all events in period
        query = AuditTrailQuery(
            start_date=start_date,
            end_date=end_date,
            limit=10000,  # High limit for full report
        )
        events = await self.query_audit_events(query)

        # Group by type
        events_by_type: dict[str, int] = {}
        users: set[str] = set()

        for event in events:
            event_type = event.event_type.value
            events_by_type[event_type] = events_by_type.get(event_type, 0) + 1
            users.add(event.user)

        return AuditReportOut(
            period_start=start_date,
            period_end=end_date,
            total_events=len(events),
            events_by_type=events_by_type,
            users=sorted(list(users)),
            events=events,
        )

    async def get_event_statistics(
        self, start_date: date, end_date: date
    ) -> dict:
        """Get audit event statistics for a period.

        Args:
            start_date: Period start.
            end_date: Period end.

        Returns:
            Statistics dictionary.
        """
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())

        # Count by event type
        type_counts = await self.db.execute(
            select(
                AuditEvent.event_type,
                func.count(AuditEvent.event_id).label("count"),
            )
            .where(AuditEvent.timestamp.between(start_datetime, end_datetime))
            .group_by(AuditEvent.event_type)
        )
        by_type = {row.event_type: row.count for row in type_counts.all()}

        # Count by user
        user_counts = await self.db.execute(
            select(
                AuditEvent.user,
                func.count(AuditEvent.event_id).label("count"),
            )
            .where(AuditEvent.timestamp.between(start_datetime, end_datetime))
            .group_by(AuditEvent.user)
        )
        by_user = {row.user: row.count for row in user_counts.all()}

        # Total count
        total_count = await self.db.execute(
            select(func.count(AuditEvent.event_id)).where(
                AuditEvent.timestamp.between(start_datetime, end_datetime)
            )
        )
        total = total_count.scalar() or 0

        return {
            "period_start": str(start_date),
            "period_end": str(end_date),
            "total_events": total,
            "by_event_type": by_type,
            "by_user": by_user,
        }

    async def export_audit_report_excel(
        self, start_date: date, end_date: date
    ) -> bytes:
        """Export audit report as Excel file.

        Args:
            start_date: Report period start.
            end_date: Report period end.

        Returns:
            Excel file as bytes.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        report = await self.generate_audit_report(start_date, end_date)

        # Create workbook
        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Add summary info
        ws_summary["A1"] = "Audit Report"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A3"] = "Period:"
        ws_summary["B3"] = f"{start_date} to {end_date}"
        ws_summary["A4"] = "Total Events:"
        ws_summary["B4"] = report.total_events
        ws_summary["A5"] = "Unique Users:"
        ws_summary["B5"] = len(report.users)

        # Events by type
        ws_summary["A7"] = "Events by Type"
        ws_summary["A7"].font = header_font
        row = 8
        for event_type, count in report.events_by_type.items():
            ws_summary[f"A{row}"] = event_type
            ws_summary[f"B{row}"] = count
            row += 1

        # Events detail sheet
        ws_events = wb.create_sheet("Events")
        headers = ["Event ID", "Type", "User", "Timestamp", "IP Address", "Details"]
        for col, header in enumerate(headers, 1):
            cell = ws_events.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, event in enumerate(report.events, 2):
            ws_events.cell(row=row_idx, column=1, value=event.event_id)
            ws_events.cell(row=row_idx, column=2, value=event.event_type.value)
            ws_events.cell(row=row_idx, column=3, value=event.user)
            ws_events.cell(row=row_idx, column=4, value=str(event.timestamp))
            ws_events.cell(row=row_idx, column=5, value=event.ip_address or "")
            ws_events.cell(row=row_idx, column=6, value=str(event.details))

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()


# Convenience function for logging events
async def log_event(
    db: AsyncSession,
    event_type: AuditEventType,
    user: str,
    details: dict,
    ip_address: Optional[str] = None,
) -> AuditEventOut:
    """Convenience function to log an audit event.

    Args:
        db: Database session.
        event_type: Type of event.
        user: User who performed the action.
        details: Event details.
        ip_address: Optional client IP.

    Returns:
        Created AuditEventOut.
    """
    audit_trail = AuditTrail(db)
    return await audit_trail.log_audit_event(event_type, user, details, ip_address)
